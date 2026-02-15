from __future__ import annotations

"""Excel (XLSX) export helpers.

This project already supports exporting tables to PDF.
This module provides the analogous XLSX export.

Design goals:
- small surface area (openpyxl only)
- safe defaults (limits handled by caller)
- output that is immediately usable in Excel: frozen headers + auto-filter
- optional "auto" chart when the result looks like (category, value)
"""

from io import BytesIO
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


def _is_number(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        s = str(v).strip().replace(" ", "")
        if not s:
            return False
        float(s.replace(",", "."))
        return True
    except Exception:
        return False


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace(" ", "")
        if not s:
            return None
        return float(s.replace(",", "."))
    except Exception:
        return None


def _pick_chart_columns(columns: Sequence[str], rows: Sequence[Sequence[Any]]) -> tuple[int, int] | None:
    """Pick (category_idx, value_idx) for an "auto" chart.

    Heuristic:
    - value_idx: column with the highest count of numeric values in a sample
    - category_idx: first column that isn't the chosen value column
    """
    if not columns or not rows:
        return None

    ncols = len(columns)
    sample = rows[:200]

    numeric_counts = [0] * ncols
    for r in sample:
        for i in range(min(ncols, len(r))):
            if _is_number(r[i]):
                numeric_counts[i] += 1

    # best numeric column must have at least a few numeric values
    value_idx = max(range(ncols), key=lambda i: numeric_counts[i])
    if numeric_counts[value_idx] < max(3, min(10, len(sample) // 5 or 1)):
        return None

    category_idx = 0 if value_idx != 0 else (1 if ncols > 1 else 0)
    if category_idx == value_idx:
        return None

    return category_idx, value_idx


def table_to_xlsx_bytes(
    title: str,
    columns: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    chart: str = "auto",  # auto|none
    chart_title: str | None = None,
    max_chart_rows: int = 25,
) -> bytes:
    """Create an XLSX file in memory.

    Args:
        title: Used for the workbook properties and chart title.
        columns/rows: tabular data.
        chart: "auto" to add a simple bar chart when possible, "none" otherwise.
        max_chart_rows: cap chart points to keep files light/readable.
    """

    safe_title = (title or "Export").strip() or "Export"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header
    ws.append([str(c) for c in (columns or [])])
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    # Data rows
    for r in rows or []:
        # keep row length aligned with columns
        out = list(r[: len(columns)])
        if len(out) < len(columns):
            out += [None] * (len(columns) - len(out))
        ws.append(out)

    # Freeze header + filter
    if columns:
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(columns))
        last_row = max(1, 1 + (len(rows) if rows else 0))
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Column widths (best-effort)
    for i, col_name in enumerate(columns or [], start=1):
        max_len = len(str(col_name))
        for r in (rows or [])[:50]:
            if i - 1 < len(r) and r[i - 1] is not None:
                max_len = max(max_len, len(str(r[i - 1])))
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(45, max_len + 2))

    # Optional chart
    if (chart or "auto").lower() == "auto":
        pick = _pick_chart_columns(list(columns or []), list(rows or []))
        if pick:
            cat_idx, val_idx = pick
            ws_chart = wb.create_sheet("Chart")

            # Decide chart range: first N rows (assume SQL already ORDER BY DESC when desired)
            n_points = min(len(rows or []), max(1, int(max_chart_rows)))
            if n_points > 0:
                # Normalize value column to numeric where possible (Excel charts behave better)
                for ridx in range(2, 2 + n_points):
                    v = ws.cell(row=ridx, column=val_idx + 1).value
                    vf = _to_float(v)
                    if vf is not None:
                        ws.cell(row=ridx, column=val_idx + 1).value = vf

                data_ref = Reference(ws, min_col=val_idx + 1, min_row=1, max_row=1 + n_points)
                cats_ref = Reference(ws, min_col=cat_idx + 1, min_row=2, max_row=1 + n_points)

                ch = BarChart()
                ch.type = "col"
                ch.title = chart_title or safe_title
                ch.y_axis.title = str(columns[val_idx])
                ch.x_axis.title = str(columns[cat_idx])
                ch.add_data(data_ref, titles_from_data=True)
                ch.set_categories(cats_ref)
                ch.height = 12
                ch.width = 26

                ws_chart.add_chart(ch, "A1")

    # Workbook properties
    try:
        wb.properties.title = safe_title[:200]
    except Exception:
        pass

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
