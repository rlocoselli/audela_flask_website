from __future__ import annotations

import csv
import os
import shutil
import subprocess
from io import BytesIO
from io import StringIO
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

from flask import Response, current_app, flash, g, redirect, render_template, request, url_for
from flask_login import current_user, login_user, logout_user
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

from ...extensions import db
from ...i18n import tr
from ...models import BillingEvent, Role, SubscriptionPlan, Tenant, TenantSubscription, User
from ...tenancy import clear_current_tenant
from . import bp


ALLOWED_BILLING_CYCLES = {"monthly", "yearly"}
CELERY_SERVICES = ("celery-worker", "celery-beat")


def _month_start(value: datetime) -> datetime:
    return datetime(value.year, value.month, 1)


def _add_months(value: datetime, months: int) -> datetime:
    month_index = (value.month - 1) + months
    year = value.year + (month_index // 12)
    month = (month_index % 12) + 1
    return datetime(year, month, 1)


def _normalize_mrr(subscription: TenantSubscription | None) -> Decimal:
    if not subscription or not subscription.plan:
        return Decimal("0.00")
    if subscription.billing_cycle == "yearly":
        return Decimal(subscription.plan.price_yearly or 0) / Decimal("12")
    return Decimal(subscription.plan.price_monthly or 0)


def _build_admin_finance(subscriptions: list[TenantSubscription], selected_months_raw: str | None) -> dict:
    try:
        selected_months_int = int(selected_months_raw or "12")
    except (TypeError, ValueError):
        selected_months_int = 12
    if selected_months_int < 1:
        selected_months_int = 1
    if selected_months_int > 36:
        selected_months_int = 36

    now = datetime.utcnow()
    period_end_month = _month_start(now)
    period_start_month = _add_months(period_end_month, -(selected_months_int - 1))
    period_end_exclusive = _add_months(period_end_month, 1)

    month_slots: list[datetime] = []
    cursor = period_start_month
    while cursor < period_end_exclusive:
        month_slots.append(cursor)
        cursor = _add_months(cursor, 1)

    monthly_report_map = {
        slot.strftime("%Y-%m"): {
            "month_key": slot.strftime("%Y-%m"),
            "month_label": slot.strftime("%m/%Y"),
            "new_subscriptions": 0,
            "cancelled_subscriptions": 0,
            "new_mrr": Decimal("0.00"),
            "realized_revenue": Decimal("0.00"),
        }
        for slot in month_slots
    }

    status_totals = defaultdict(int)
    current_mrr = Decimal("0.00")
    period_new_subscriptions = 0
    period_cancelled_subscriptions = 0
    plan_mix_map: dict[str, dict] = {}
    currency_totals = defaultdict(int)

    for sub in subscriptions:
        status_key = (sub.status or "unknown").strip().lower()
        status_totals[status_key] += 1

        if sub.plan and sub.plan.currency:
            currency_totals[sub.plan.currency] += 1

        if status_key in {"active", "trial"} and sub.plan:
            sub_mrr = _normalize_mrr(sub)
            current_mrr += sub_mrr

            plan_key = sub.plan.code
            if plan_key not in plan_mix_map:
                plan_mix_map[plan_key] = {
                    "plan_code": sub.plan.code,
                    "plan_name": sub.plan.name,
                    "subscriptions": 0,
                    "mrr": Decimal("0.00"),
                }
            plan_mix_map[plan_key]["subscriptions"] += 1
            plan_mix_map[plan_key]["mrr"] += sub_mrr

        created_ref = sub.created_at or sub.current_period_start or sub.trial_start_date
        if created_ref and period_start_month <= created_ref < period_end_exclusive:
            period_new_subscriptions += 1
            created_key = created_ref.strftime("%Y-%m")
            row = monthly_report_map.get(created_key)
            if row is not None:
                row["new_subscriptions"] += 1
                row["new_mrr"] += _normalize_mrr(sub)

        if sub.cancelled_at and period_start_month <= sub.cancelled_at < period_end_exclusive:
            period_cancelled_subscriptions += 1
            cancelled_key = sub.cancelled_at.strftime("%Y-%m")
            row = monthly_report_map.get(cancelled_key)
            if row is not None:
                row["cancelled_subscriptions"] += 1

    period_events = (
        BillingEvent.query
        .filter(
            BillingEvent.created_at >= period_start_month,
            BillingEvent.created_at < period_end_exclusive,
            BillingEvent.amount.isnot(None),
        )
        .all()
    )
    realized_revenue_period = Decimal("0.00")
    for event in period_events:
        amount = Decimal(event.amount or 0)
        if amount <= 0:
            continue
        realized_revenue_period += amount
        if event.currency:
            currency_totals[event.currency] += 1
        event_key = event.created_at.strftime("%Y-%m") if event.created_at else None
        row = monthly_report_map.get(event_key) if event_key else None
        if row is not None:
            row["realized_revenue"] += amount

    default_currency = "EUR"
    if currency_totals:
        default_currency = sorted(currency_totals.items(), key=lambda item: item[1], reverse=True)[0][0]

    monthly_subscription_report = list(monthly_report_map.values())
    for row in monthly_subscription_report:
        row["net_subscriptions"] = row["new_subscriptions"] - row["cancelled_subscriptions"]

    plan_mix = sorted(
        plan_mix_map.values(),
        key=lambda item: (item["mrr"], item["subscriptions"]),
        reverse=True,
    )

    return {
        "selected_months": selected_months_int,
        "currency": default_currency,
        "period_start": period_start_month,
        "period_end": _add_months(period_end_exclusive, -1),
        "current_mrr": current_mrr,
        "arr_proxy": current_mrr * Decimal("12"),
        "realized_revenue_period": realized_revenue_period,
        "period_new_subscriptions": period_new_subscriptions,
        "period_cancelled_subscriptions": period_cancelled_subscriptions,
        "active_subscriptions": status_totals.get("active", 0),
        "trial_subscriptions": status_totals.get("trial", 0),
        "suspended_subscriptions": status_totals.get("suspended", 0),
        "cancelled_total": status_totals.get("cancelled", 0),
        "monthly_report": monthly_subscription_report,
        "plan_mix": plan_mix,
    }


def _is_platform_admin(user: User | None) -> bool:
    return bool(user and user.is_authenticated and user.has_role("platform_admin"))


def _admin_guard_redirect():
    if not current_user.is_authenticated:
        return redirect(url_for("admin.login"))
    if not current_user.has_role("platform_admin"):
        flash(tr("Acesso negado.", getattr(g, "lang", None)), "error")
        return redirect(url_for("portal.home"))
    return None


def _project_root() -> str:
    return os.path.abspath(os.path.join(current_app.root_path, os.pardir))


def _truncate_text(value: str, limit: int = 6000) -> str:
    if len(value) <= limit:
        return value
    return f"{value[:limit]}\n... (truncated)"


def _resolve_compose_command() -> list[str] | None:
    docker_path = shutil.which("docker")
    if docker_path:
        try:
            probe = subprocess.run(
                [docker_path, "compose", "version"],
                capture_output=True,
                text=True,
                check=False,
                timeout=8,
            )
            if probe.returncode == 0:
                return [docker_path, "compose"]
        except Exception:
            pass

    docker_compose_path = shutil.which("docker-compose")
    if docker_compose_path:
        return [docker_compose_path]

    return None


def _compose_unavailable_reason() -> str | None:
    if _resolve_compose_command() is not None:
        return None

    if os.path.exists("/.dockerenv"):
        return (
            "Docker Compose is not available inside the web app container runtime. "
            "Use your external server workflow to start/stop Celery services."
        )

    return "Docker Compose command not available on this host."


def _compose_file_args(project_root: str) -> list[str]:
    compose_args: list[str] = []
    for filename in ("docker-compose.yml", "docker-compose.letsencrypt.yml"):
        abs_path = os.path.join(project_root, filename)
        if os.path.exists(abs_path):
            compose_args.extend(["-f", filename])
    return compose_args


def _run_compose_command(action_label: str, command: list[str], project_root: str, timeout: int = 90) -> dict:
    try:
        proc = subprocess.run(
            command,
            cwd=project_root,
            capture_output=True,
            text=True,
            check=False,
            timeout=timeout,
        )
        return {
            "step": action_label,
            "ok": proc.returncode == 0,
            "returncode": proc.returncode,
            "command": " ".join(command),
            "stdout": _truncate_text(proc.stdout or ""),
            "stderr": _truncate_text(proc.stderr or ""),
        }
    except subprocess.TimeoutExpired:
        return {
            "step": action_label,
            "ok": False,
            "returncode": None,
            "command": " ".join(command),
            "stdout": "",
            "stderr": f"Command timed out after {timeout}s.",
        }
    except Exception as exc:
        return {
            "step": action_label,
            "ok": False,
            "returncode": None,
            "command": " ".join(command),
            "stdout": "",
            "stderr": str(exc),
        }


def _run_celery_service_action(action: str) -> dict:
    compose_cmd = _resolve_compose_command()
    if not compose_cmd:
        return {
            "ok": False,
            "action": action,
            "error": _compose_unavailable_reason() or "Docker Compose command not available on this host.",
            "steps": [],
        }

    project_root = _project_root()
    compose_args = _compose_file_args(project_root)
    base = [*compose_cmd, *compose_args]

    plans: dict[str, list[tuple[str, list[str], int]]] = {
        "enable": [
            ("start", [*base, "up", "-d", "--no-deps", *CELERY_SERVICES], 120),
            ("status", [*base, "ps", *CELERY_SERVICES], 45),
            ("logs", [*base, "logs", "--tail=60", *CELERY_SERVICES], 60),
        ],
        "disable": [
            ("stop", [*base, "stop", *CELERY_SERVICES], 60),
            ("remove", [*base, "rm", "-f", *CELERY_SERVICES], 60),
            ("status", [*base, "ps", *CELERY_SERVICES], 45),
        ],
        "restart": [
            ("ensure-up", [*base, "up", "-d", "--no-deps", *CELERY_SERVICES], 120),
            ("restart", [*base, "restart", *CELERY_SERVICES], 90),
            ("status", [*base, "ps", *CELERY_SERVICES], 45),
            ("logs", [*base, "logs", "--tail=60", *CELERY_SERVICES], 60),
        ],
        "status": [
            ("status", [*base, "ps", *CELERY_SERVICES], 45),
            ("logs", [*base, "logs", "--tail=60", *CELERY_SERVICES], 60),
        ],
    }

    if action not in plans:
        return {
            "ok": False,
            "action": action,
            "error": "Unknown action.",
            "steps": [],
        }

    steps: list[dict] = []
    overall_ok = True
    for label, command, timeout in plans[action]:
        step_result = _run_compose_command(label, command, project_root, timeout=timeout)
        steps.append(step_result)
        if not step_result["ok"]:
            overall_ok = False

    return {
        "ok": overall_ok,
        "action": action,
        "error": None,
        "steps": steps,
        "compose_command": " ".join(compose_cmd),
        "compose_files": compose_args,
    }


def _build_celery_screen_context() -> dict:
    from audela.celery_app import celery_app

    broker = current_app.config.get("CELERY_BROKER_URL") or current_app.config.get("REDIS_URL")
    backend = current_app.config.get("CELERY_RESULT_BACKEND") or broker

    worker_ping: dict | None = None
    worker_ping_error: str | None = None
    try:
        # Use a slightly longer timeout to reduce false negatives on busy hosts.
        worker_ping = celery_app.control.inspect(timeout=3.0).ping()
        if not worker_ping:
            worker_ping_error = "No worker replied to inspect ping."
    except Exception as exc:
        worker_ping_error = str(exc)

    compose_reason = _compose_unavailable_reason()

    return {
        "broker_url": broker,
        "result_backend": backend,
        "default_queue": current_app.config.get("CELERY_DEFAULT_QUEUE", "default"),
        "task_ignore_result": bool(current_app.config.get("CELERY_TASK_IGNORE_RESULT", False)),
        "worker_ping": worker_ping,
        "worker_ping_error": worker_ping_error,
        "compose_available": compose_reason is None,
        "compose_unavailable_reason": compose_reason,
    }


def _handle_celery_intent(intent: str) -> tuple[dict | None, dict | None, dict | None]:
    service_action_result = None
    task_result = None
    task_lookup = None

    if intent == "service_action":
        action = (request.form.get("action") or "status").strip().lower()
        service_action_result = _run_celery_service_action(action)
        if service_action_result["ok"]:
            flash(
                tr("Celery action completed: {action}", getattr(g, "lang", None), action=action),
                "success",
            )
        else:
            flash(
                tr("Celery action failed: {action}", getattr(g, "lang", None), action=action),
                "error",
            )

    elif intent == "test_task":
        from celery.exceptions import TimeoutError as CeleryTimeoutError
        from audela.celery_app import celery_app

        wait_raw = (request.form.get("wait_seconds") or "8").strip()
        try:
            wait_seconds = max(1, min(int(wait_raw), 30))
        except ValueError:
            wait_seconds = 8

        try:
            job = celery_app.send_task("audela.tasks.celery_healthcheck")
            task_result = {
                "task_id": job.id,
                "state": "SENT",
                "wait_seconds": wait_seconds,
                "result": None,
                "error": None,
            }
            try:
                payload = job.get(timeout=wait_seconds)
                task_result["state"] = "SUCCESS"
                task_result["result"] = payload
                flash(tr("Celery test succeeded.", getattr(g, "lang", None)), "success")
            except CeleryTimeoutError as exc:
                task_result["state"] = "PENDING"
                task_result["error"] = str(exc)
                if bool(current_app.config.get("CELERY_TASK_IGNORE_RESULT", False)):
                    flash(
                        tr(
                            "Celery task sent. Result backend is disabled, so completion cannot be confirmed from the web app.",
                            getattr(g, "lang", None),
                        ),
                        "warning",
                    )
                else:
                    flash(tr("Celery task sent but no worker response yet.", getattr(g, "lang", None)), "warning")
            except Exception as exc:
                task_result["state"] = "FAILURE"
                task_result["error"] = str(exc)
                flash(tr("Celery task failed during execution.", getattr(g, "lang", None)), "error")
        except Exception as exc:
            task_result = {
                "task_id": None,
                "state": "ERROR",
                "wait_seconds": wait_seconds,
                "result": None,
                "error": str(exc),
            }
            flash(tr("Failed to send Celery test task.", getattr(g, "lang", None)), "error")

    elif intent == "task_lookup":
        from audela.celery_app import celery_app

        task_id = (request.form.get("task_id") or "").strip()
        if not task_id:
            flash(tr("Provide a task id.", getattr(g, "lang", None)), "error")
        else:
            async_result = celery_app.AsyncResult(task_id)
            task_lookup = {
                "task_id": task_id,
                "state": async_result.state,
                "result": async_result.result if async_result.state == "SUCCESS" else None,
                "error": str(async_result.result) if async_result.state in {"FAILURE", "REVOKED"} else None,
            }

    return service_action_result, task_result, task_lookup


@bp.route("/celery", methods=["GET", "POST"])
def celery_console():
    guard = _admin_guard_redirect()
    if guard is not None:
        return guard

    service_action_result = None
    task_result = None
    task_lookup = None

    if request.method == "POST":
        intent = (request.form.get("intent") or "").strip().lower()
        service_action_result, task_result, task_lookup = _handle_celery_intent(intent)

        # Dashboard quick-actions post here, then bounce back to /admin.
        if (request.form.get("return_to") or "").strip().lower() == "dashboard":
            selected_months = (request.form.get("months") or request.args.get("months") or "12").strip()
            return redirect(url_for("admin.dashboard", months=selected_months))

    context = _build_celery_screen_context()
    return render_template(
        "admin/celery.html",
        service_action_result=service_action_result,
        task_result=task_result,
        task_lookup=task_lookup,
        **context,
    )


@bp.route("/login", methods=["GET", "POST"])
def login():
    if _is_platform_admin(current_user):
        return redirect(url_for("admin.dashboard"))

    if request.method == "POST":
        identifier = (request.form.get("identifier") or "").strip()
        password = request.form.get("password") or ""

        if not identifier or not password:
            flash(tr("Preencha todos os campos.", getattr(g, "lang", None)), "error")
            return render_template("admin/login.html")

        user = (
            User.query
            .join(User.roles)
            .filter(Role.code == "platform_admin", User.email == identifier)
            .first()
        )

        if not user or not user.check_password(password):
            flash(tr("Credenciais inválidas.", getattr(g, "lang", None)), "error")
            return render_template("admin/login.html")

        login_user(user)
        clear_current_tenant()
        user.last_login_at = datetime.utcnow()
        db.session.commit()

        flash(tr("Login admin realizado com sucesso.", getattr(g, "lang", None)), "success")
        return redirect(url_for("admin.dashboard"))

    return render_template("admin/login.html")


@bp.route("/logout")
def logout():
    logout_user()
    clear_current_tenant()
    return redirect(url_for("admin.login"))


@bp.route("/", methods=["GET", "POST"])
def dashboard():
    guard = _admin_guard_redirect()
    if guard is not None:
        return guard

    if request.method == "POST":
        intent = (request.form.get("intent") or "").strip().lower()
        if intent in {"service_action", "test_task"}:
            _handle_celery_intent(intent)
        selected_months = (request.form.get("months") or request.args.get("months") or "12").strip()
        return redirect(url_for("admin.dashboard", months=selected_months))

    plans = (
        SubscriptionPlan.query
        .filter_by(is_active=True)
        .order_by(SubscriptionPlan.display_order.asc(), SubscriptionPlan.name.asc())
        .all()
    )
    tenants = Tenant.query.order_by(Tenant.created_at.desc()).all()
    users = User.query.order_by(User.created_at.desc()).all()
    subscriptions = TenantSubscription.query.all()

    subscription_by_tenant = {row.tenant_id: row for row in subscriptions}
    tenant_by_id = {tenant.id: tenant for tenant in tenants}

    users_view = []
    for user in users:
        tenant = tenant_by_id.get(user.tenant_id)
        subscription = subscription_by_tenant.get(user.tenant_id)
        users_view.append(
            {
                "user": user,
                "tenant": tenant,
                "subscription": subscription,
                "plan_code": subscription.plan.code if subscription and subscription.plan else (tenant.plan if tenant else "-"),
                "plan_name": subscription.plan.name if subscription and subscription.plan else (tenant.plan if tenant else "-"),
            }
        )
    admin_finance = _build_admin_finance(subscriptions, request.args.get("months", "12"))
    celery_overview = _build_celery_screen_context()

    return render_template(
        "admin/dashboard.html",
        plans=plans,
        tenants=tenants,
        users_view=users_view,
        subscriptions=subscriptions,
        admin_finance=admin_finance,
        celery_overview=celery_overview,
    )


@bp.route("/subscription-report.csv")
def export_subscription_report_csv():
    guard = _admin_guard_redirect()
    if guard is not None:
        return guard

    subscriptions = TenantSubscription.query.all()
    admin_finance = _build_admin_finance(subscriptions, request.args.get("months", "12"))

    csv_buffer = StringIO()
    writer = csv.writer(csv_buffer)

    writer.writerow(["report", "subscription_financial_control"])
    writer.writerow(["period_months", admin_finance["selected_months"]])
    writer.writerow(["period_start", admin_finance["period_start"].strftime("%Y-%m")])
    writer.writerow(["period_end", admin_finance["period_end"].strftime("%Y-%m")])
    writer.writerow(["currency", admin_finance["currency"]])
    writer.writerow([])

    writer.writerow(["kpi", "value"])
    writer.writerow(["current_mrr", f"{admin_finance['current_mrr']:.2f}"])
    writer.writerow(["arr_proxy", f"{admin_finance['arr_proxy']:.2f}"])
    writer.writerow(["realized_revenue_period", f"{admin_finance['realized_revenue_period']:.2f}"])
    writer.writerow(["active_subscriptions", admin_finance["active_subscriptions"]])
    writer.writerow(["trial_subscriptions", admin_finance["trial_subscriptions"]])
    writer.writerow(["suspended_subscriptions", admin_finance["suspended_subscriptions"]])
    writer.writerow(["period_new_subscriptions", admin_finance["period_new_subscriptions"]])
    writer.writerow(["period_cancelled_subscriptions", admin_finance["period_cancelled_subscriptions"]])
    writer.writerow([])

    writer.writerow(["monthly_report"])
    writer.writerow(["month", "new_subscriptions", "cancelled_subscriptions", "net_subscriptions", "new_mrr", "realized_revenue"])
    for row in admin_finance["monthly_report"]:
        writer.writerow([
            row["month_key"],
            row["new_subscriptions"],
            row["cancelled_subscriptions"],
            row["net_subscriptions"],
            f"{row['new_mrr']:.2f}",
            f"{row['realized_revenue']:.2f}",
        ])
    writer.writerow([])

    writer.writerow(["plan_mix"])
    writer.writerow(["plan_code", "plan_name", "subscriptions", "mrr"])
    for row in admin_finance["plan_mix"]:
        writer.writerow([
            row["plan_code"],
            row["plan_name"],
            row["subscriptions"],
            f"{row['mrr']:.2f}",
        ])

    filename = f"admin_subscription_report_{admin_finance['period_start'].strftime('%Y%m')}_{admin_finance['period_end'].strftime('%Y%m')}.csv"
    return Response(
        csv_buffer.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@bp.route("/subscription-report.xlsx")
def export_subscription_report_xlsx():
    guard = _admin_guard_redirect()
    if guard is not None:
        return guard

    subscriptions = TenantSubscription.query.all()
    admin_finance = _build_admin_finance(subscriptions, request.args.get("months", "12"))

    wb = Workbook()

    ws_kpi = wb.active
    ws_kpi.title = "KPI"
    ws_kpi.append(["Metric", "Value"])
    ws_kpi.append(["Period months", admin_finance["selected_months"]])
    ws_kpi.append(["Period start", admin_finance["period_start"].strftime("%Y-%m")])
    ws_kpi.append(["Period end", admin_finance["period_end"].strftime("%Y-%m")])
    ws_kpi.append(["Currency", admin_finance["currency"]])
    ws_kpi.append(["Current MRR", float(admin_finance["current_mrr"])])
    ws_kpi.append(["ARR proxy", float(admin_finance["arr_proxy"])])
    ws_kpi.append(["Realized revenue period", float(admin_finance["realized_revenue_period"])])
    ws_kpi.append(["Active subscriptions", admin_finance["active_subscriptions"]])
    ws_kpi.append(["Trial subscriptions", admin_finance["trial_subscriptions"]])
    ws_kpi.append(["Suspended subscriptions", admin_finance["suspended_subscriptions"]])
    ws_kpi.append(["New subscriptions period", admin_finance["period_new_subscriptions"]])
    ws_kpi.append(["Cancelled subscriptions period", admin_finance["period_cancelled_subscriptions"]])
    ws_kpi.column_dimensions["A"].width = 32
    ws_kpi.column_dimensions["B"].width = 24

    ws_monthly = wb.create_sheet(title="Mensuel")
    ws_monthly.append(["Month", "New subscriptions", "Cancelled subscriptions", "Net subscriptions", "New MRR", "Realized revenue"])
    for row in admin_finance["monthly_report"]:
        ws_monthly.append([
            row["month_key"],
            row["new_subscriptions"],
            row["cancelled_subscriptions"],
            row["net_subscriptions"],
            float(row["new_mrr"]),
            float(row["realized_revenue"]),
        ])
    ws_monthly.column_dimensions["A"].width = 14
    ws_monthly.column_dimensions["B"].width = 18
    ws_monthly.column_dimensions["C"].width = 22
    ws_monthly.column_dimensions["D"].width = 16
    ws_monthly.column_dimensions["E"].width = 16
    ws_monthly.column_dimensions["F"].width = 16

    ws_plans = wb.create_sheet(title="Plans")
    ws_plans.append(["Plan code", "Plan name", "Subscriptions", "MRR"])
    for row in admin_finance["plan_mix"]:
        ws_plans.append([
            row["plan_code"],
            row["plan_name"],
            row["subscriptions"],
            float(row["mrr"]),
        ])
    ws_plans.column_dimensions["A"].width = 20
    ws_plans.column_dimensions["B"].width = 28
    ws_plans.column_dimensions["C"].width = 14
    ws_plans.column_dimensions["D"].width = 14

    ws_summary = wb.create_sheet(title="Résumé")
    ws_summary.append(["KPI", "Valeur"])
    ws_summary.append(["Devise", admin_finance["currency"]])
    ws_summary.append(["MRR actuel", float(admin_finance["current_mrr"])])
    ws_summary.append(["ARR estimé", float(admin_finance["arr_proxy"])])
    ws_summary.append(["Revenu réalisé (période)", float(admin_finance["realized_revenue_period"])])
    ws_summary.append(["Actifs", admin_finance["active_subscriptions"]])
    ws_summary.append(["Trials", admin_finance["trial_subscriptions"]])
    ws_summary.append(["Suspendus", admin_finance["suspended_subscriptions"]])
    ws_summary.append([])
    ws_summary.append(["Mois", "Nouveaux abonnements", "Revenu réalisé"])

    for row in admin_finance["monthly_report"]:
        ws_summary.append([
            row["month_key"],
            row["new_subscriptions"],
            float(row["realized_revenue"]),
        ])

    ws_summary.column_dimensions["A"].width = 26
    ws_summary.column_dimensions["B"].width = 24
    ws_summary.column_dimensions["C"].width = 20

    chart_start_row = 11
    chart_end_row = chart_start_row + len(admin_finance["monthly_report"]) - 1
    if chart_end_row >= chart_start_row:
        revenue_chart = BarChart()
        revenue_chart.title = "Évolution mensuelle"
        revenue_chart.y_axis.title = admin_finance["currency"]
        revenue_chart.x_axis.title = "Mois"
        revenue_chart.height = 8
        revenue_chart.width = 16

        revenue_values = Reference(ws_summary, min_col=3, min_row=10, max_col=3, max_row=chart_end_row)
        categories = Reference(ws_summary, min_col=1, min_row=11, max_row=chart_end_row)
        revenue_chart.add_data(revenue_values, titles_from_data=True)
        revenue_chart.set_categories(categories)
        revenue_chart.style = 10

        subs_line = LineChart()
        subs_line.y_axis.title = "Abonnements"
        subs_line.y_axis.axId = 200
        subs_line.y_axis.crosses = "max"
        subs_line_values = Reference(ws_summary, min_col=2, min_row=10, max_col=2, max_row=chart_end_row)
        subs_line.add_data(subs_line_values, titles_from_data=True)
        subs_line.set_categories(categories)

        revenue_chart += subs_line
        ws_summary.add_chart(revenue_chart, "E2")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"admin_subscription_report_{admin_finance['period_start'].strftime('%Y%m')}_{admin_finance['period_end'].strftime('%Y%m')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@bp.route("/set-user-plan", methods=["POST"])
def set_user_plan():
    guard = _admin_guard_redirect()
    if guard is not None:
        return guard

    user_id_raw = request.form.get("user_id")
    plan_code = (request.form.get("plan_code") or "").strip()
    billing_cycle = (request.form.get("billing_cycle") or "monthly").strip().lower()

    if not user_id_raw or not plan_code:
        flash(tr("Preencha todos os campos.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    if billing_cycle not in ALLOWED_BILLING_CYCLES:
        billing_cycle = "monthly"

    try:
        user_id = int(user_id_raw)
    except ValueError:
        flash(tr("Configuration inválida.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    user = User.query.get(user_id)
    plan = SubscriptionPlan.query.filter_by(code=plan_code, is_active=True).first()
    if not user or not plan:
        flash(tr("Configuração inválida.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    tenant = Tenant.query.get(user.tenant_id)
    if not tenant:
        flash(tr("Tenant not found", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    now = datetime.utcnow()
    sub = TenantSubscription.query.filter_by(tenant_id=tenant.id).first()
    if not sub:
        sub = TenantSubscription(
            tenant_id=tenant.id,
            plan_id=plan.id,
            status="active",
            billing_cycle=billing_cycle,
            current_period_start=now,
            current_period_end=now,
            next_billing_date=now,
            current_users_count=0,
            current_companies_count=0,
            transactions_this_month=0,
        )
        db.session.add(sub)

    sub.plan_id = plan.id
    sub.status = "active"
    sub.billing_cycle = billing_cycle
    sub.current_period_start = now
    if billing_cycle == "yearly":
        sub.current_period_end = datetime(now.year + 1, now.month, min(now.day, 28), now.hour, now.minute, now.second)
    else:
        next_month = now.month + 1
        year = now.year
        if next_month > 12:
            year += 1
            next_month = 1
        sub.current_period_end = datetime(year, next_month, min(now.day, 28), now.hour, now.minute, now.second)
    sub.next_billing_date = sub.current_period_end

    tenant.plan = plan.code

    db.session.add(
        BillingEvent(
            tenant_id=tenant.id,
            subscription_id=sub.id,
            event_type="admin.subscription.updated",
            amount=plan.price_yearly if billing_cycle == "yearly" else plan.price_monthly,
            currency=plan.currency,
            metadata_json={
                "updated_by_user_id": current_user.id,
                "target_user_id": user.id,
                "plan_code": plan.code,
                "billing_cycle": billing_cycle,
            },
        )
    )
    db.session.commit()

    flash(
        tr("Assinatura atualizada para {email}.", getattr(g, "lang", None), email=user.email),
        "success",
    )
    return redirect(url_for("admin.dashboard"))


@bp.route("/change-password", methods=["POST"])
def change_admin_password():
    guard = _admin_guard_redirect()
    if guard is not None:
        return guard

    current_password = request.form.get("current_password") or ""
    new_password = request.form.get("new_password") or ""
    confirm_password = request.form.get("confirm_password") or ""

    if not current_password or not new_password or not confirm_password:
        flash(tr("Preencha todos os campos.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    if not current_user.check_password(current_password):
        flash(tr("Senha atual inválida.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    if new_password != confirm_password:
        flash(tr("As senhas não coincidem.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    if len(new_password) < 12:
        flash(tr("A senha deve ter pelo menos 12 caracteres.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    if new_password == current_password:
        flash(tr("A nova senha deve ser diferente da senha atual.", getattr(g, "lang", None)), "error")
        return redirect(url_for("admin.dashboard"))

    current_user.set_password(new_password)
    db.session.commit()
    flash(tr("Senha alterada com sucesso.", getattr(g, "lang", None)), "success")
    return redirect(url_for("admin.dashboard"))
