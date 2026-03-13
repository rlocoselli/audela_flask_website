from __future__ import annotations

"""Excel (XLSX) export helpers.

This module is intentionally dependency-light (openpyxl only).

It produces:
- a "Data" sheet with headers, frozen panes, and auto-filter
- optionally a "Chart" sheet with a simple bar chart when the dataset
  looks like (category, numeric)
"""

from io import BytesIO
from typing import Any, Iterable
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


_TEMPLATE_STYLES: dict[str, dict[str, Any]] = {
    "clean": {
        "header_fill": "1F4E78",
        "header_font": "FFFFFF",
        "stripe_fill": "EEF5FC",
        "title_fill": None,
        "title_font": None,
        "chart_style": 10,
    },
    "modern": {
        "header_fill": "0E7490",
        "header_font": "FFFFFF",
        "stripe_fill": "ECFEFF",
        "title_fill": "0E7490",
        "title_font": "FFFFFF",
        "chart_style": 11,
    },
    "executive": {
        "header_fill": "374151",
        "header_font": "FFFFFF",
        "stripe_fill": "F3F4F6",
        "title_fill": "111827",
        "title_font": "FFFFFF",
        "chart_style": 2,
    },
    "dark": {
        "header_fill": "111827",
        "header_font": "F9FAFB",
        "stripe_fill": "E5E7EB",
        "title_fill": "030712",
        "title_font": "F9FAFB",
        "chart_style": 13,
    },
}

_COLOR_THEME_MAP = {
    "blue": "1F4E78",
    "emerald": "1E7D4D",
    "amber": "B26A00",
    "slate": "374151",
    "rose": "9F1239",
}


def _normalize_hex_color(value: str | None) -> str | None:
    if not value:
        return None
    color = str(value).strip().lstrip("#").upper()
    if re.fullmatch(r"[0-9A-F]{6}", color):
        return color
    return None


def _resolve_theme_color(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip().lower()
    if raw in _COLOR_THEME_MAP:
        return _COLOR_THEME_MAP[raw]
    return _normalize_hex_color(raw)


def _is_number(v: Any) -> bool:
    try:
        if v is None:
            return False
        float(v)
        return True
    except Exception:
        return False


def _safe_cell(v: Any) -> Any:
    # openpyxl handles numbers, strings, dates. For safety, stringify complex types.
    if v is None:
        return None
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def _to_float(v: Any) -> float | None:
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def _is_numeric_column(rows: list[list[Any]], idx: int) -> bool:
    sample = rows[: min(400, len(rows))]
    non_null = 0
    numeric = 0
    for row in sample:
        if idx >= len(row):
            continue
        v = row[idx]
        if v in (None, ""):
            continue
        non_null += 1
        if _to_float(v) is not None:
            numeric += 1
    if non_null == 0:
        return False
    return (numeric / non_null) >= 0.7


def _normalize_pivot_config(pivot_config: dict[str, Any] | None) -> dict[str, Any]:
    if not isinstance(pivot_config, dict):
        return {}

    def _split_csv(value: Any) -> list[str]:
        if isinstance(value, list):
            out = []
            for item in value:
                txt = str(item or "").strip()
                if txt:
                    out.append(txt)
            return out
        txt = str(value or "").strip()
        if not txt:
            return []
        return [p.strip() for p in txt.split(",") if p.strip()]

    rows = _split_csv(pivot_config.get("rows"))
    cols = _split_csv(pivot_config.get("columns"))
    value = str(pivot_config.get("value") or "").strip()
    agg = str(pivot_config.get("agg") or "sum").strip().lower()
    if agg not in {"sum", "count", "avg", "min", "max"}:
        agg = "sum"

    return {
        "rows": rows,
        "columns": cols,
        "value": value,
        "agg": agg,
    }


def _find_col_idx(columns: list[str], target: str) -> int:
    if not target:
        return -1
    wanted = str(target).strip().lower()
    for idx, col in enumerate(columns):
        if str(col).strip().lower() == wanted:
            return idx
    return -1


def _agg_value(current: dict[str, Any], metric: float | None, agg: str) -> None:
    current["count"] = int(current.get("count", 0)) + 1
    if metric is None:
        return
    current["sum"] = float(current.get("sum", 0.0)) + float(metric)
    if current.get("min") is None or metric < float(current.get("min")):
        current["min"] = float(metric)
    if current.get("max") is None or metric > float(current.get("max")):
        current["max"] = float(metric)


def _agg_result(state: dict[str, Any], agg: str) -> float:
    if agg == "count":
        return float(int(state.get("count", 0)))
    if agg == "avg":
        cnt = int(state.get("count", 0))
        return float(state.get("sum", 0.0)) / cnt if cnt > 0 else 0.0
    if agg == "min":
        return float(state.get("min", 0.0) if state.get("min") is not None else 0.0)
    if agg == "max":
        return float(state.get("max", 0.0) if state.get("max") is not None else 0.0)
    return float(state.get("sum", 0.0))


def _add_pivot_sheet(
    wb: Workbook,
    title: str,
    columns: list[str],
    rows: list[list[Any]],
    *,
    header_fill_color: str,
    header_font_color: str,
    stripe_fill_color: str | None,
    pivot_config: dict[str, Any] | None = None,
) -> None:
    if not columns or not rows:
        return

    cfg = _normalize_pivot_config(pivot_config)

    # Auto fallback when user didn't pick explicit pivot fields.
    numeric_idxs = [idx for idx in range(len(columns)) if _is_numeric_column(rows, idx)]
    text_idxs = [idx for idx in range(len(columns)) if idx not in numeric_idxs]

    row_fields = cfg.get("rows") or []
    col_fields = cfg.get("columns") or []
    row_idxs = [_find_col_idx(columns, f) for f in row_fields]
    row_idxs = [i for i in row_idxs if i >= 0]
    col_idx = _find_col_idx(columns, col_fields[0]) if col_fields else -1
    value_idx = _find_col_idx(columns, str(cfg.get("value") or ""))
    agg = str(cfg.get("agg") or "sum")

    if value_idx < 0:
        value_idx = numeric_idxs[0] if numeric_idxs else -1
    if not row_idxs:
        row_idxs = [text_idxs[0] if text_idxs else 0]
    if col_idx < 0:
        col_idx = text_idxs[1] if len(text_idxs) > 1 and text_idxs[1] not in row_idxs else -1

    if value_idx < 0 and agg != "count":
        return

    ws = wb.create_sheet(title="Pivot")
    ws["A1"].value = f"{(title or 'Export').strip() or 'Export'} - Pivot"
    ws["A1"].font = Font(size=14, bold=True)

    header_row = 3
    stripe_fill = PatternFill(fill_type="solid", fgColor=stripe_fill_color) if stripe_fill_color else None

    if col_idx >= 0:
        row_keys: list[str] = []
        col_keys: list[str] = []
        matrix: dict[tuple[str, str], dict[str, Any]] = {}
        row_totals: dict[str, dict[str, Any]] = {}
        col_totals: dict[str, dict[str, Any]] = {}
        grand_total: dict[str, Any] = {}

        for r in rows:
            if col_idx >= len(r):
                continue
            metric = _to_float(r[value_idx]) if value_idx >= 0 and value_idx < len(r) else None
            if agg != "count" and metric is None:
                continue
            rk_parts = []
            for ridx in row_idxs:
                rk_parts.append(str(r[ridx] if ridx < len(r) and r[ridx] not in (None, "") else "(empty)"))
            rk = " | ".join(rk_parts)
            ck = str(r[col_idx] if r[col_idx] not in (None, "") else "(empty)")
            if rk not in row_keys:
                row_keys.append(rk)
            if ck not in col_keys:
                col_keys.append(ck)
            key = (rk, ck)
            if key not in matrix:
                matrix[key] = {}
            if rk not in row_totals:
                row_totals[rk] = {}
            if ck not in col_totals:
                col_totals[ck] = {}
            _agg_value(matrix[key], metric, agg)
            _agg_value(row_totals[rk], metric, agg)
            _agg_value(col_totals[ck], metric, agg)
            _agg_value(grand_total, metric, agg)

        if not row_keys or not col_keys:
            ws["A3"].value = "No pivotable rows found"
            return

        row_label = " / ".join([str(columns[i]) for i in row_idxs])
        ws.cell(row=header_row, column=1).value = row_label
        for i, ck in enumerate(col_keys, start=2):
            ws.cell(row=header_row, column=i).value = ck
        ws.cell(row=header_row, column=2 + len(col_keys)).value = "Total"

        for c in range(1, 3 + len(col_keys)):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(bold=True, color=header_font_color)
            cell.fill = PatternFill(fill_type="solid", fgColor=header_fill_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        start_data_row = header_row + 1
        for ridx, rk in enumerate(row_keys, start=start_data_row):
            ws.cell(row=ridx, column=1).value = rk
            for cidx, ck in enumerate(col_keys, start=2):
                ws.cell(row=ridx, column=cidx).value = float(_agg_result(matrix.get((rk, ck), {}), agg))
            ws.cell(row=ridx, column=2 + len(col_keys)).value = float(_agg_result(row_totals.get(rk, {}), agg))
            if stripe_fill is not None and ((ridx - start_data_row) % 2 == 1):
                for c in range(1, 3 + len(col_keys)):
                    ws.cell(row=ridx, column=c).fill = stripe_fill

        total_row = start_data_row + len(row_keys)
        ws.cell(row=total_row, column=1).value = "Total"
        for cidx, ck in enumerate(col_keys, start=2):
            ws.cell(row=total_row, column=cidx).value = float(_agg_result(col_totals.get(ck, {}), agg))
        ws.cell(row=total_row, column=2 + len(col_keys)).value = float(_agg_result(grand_total, agg))
        for c in range(1, 3 + len(col_keys)):
            cell = ws.cell(row=total_row, column=c)
            cell.font = Font(bold=True)

        ws.freeze_panes = ws.cell(row=start_data_row, column=2)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(2 + len(col_keys))}{total_row}"
    else:
        label = " / ".join([str(columns[i]) for i in row_idxs])
        metric_label = str(columns[value_idx]) if value_idx >= 0 else "Count"
        totals: dict[str, dict[str, Any]] = {}
        for r in rows:
            if not row_idxs:
                continue
            metric = _to_float(r[value_idx]) if value_idx >= 0 and value_idx < len(r) else None
            if agg != "count" and metric is None:
                continue
            rk_parts = []
            for ridx in row_idxs:
                rk_parts.append(str(r[ridx] if ridx < len(r) and r[ridx] not in (None, "") else "(empty)"))
            rk = " | ".join(rk_parts)
            if rk not in totals:
                totals[rk] = {}
            _agg_value(totals[rk], metric, agg)

        ws.cell(row=header_row, column=1).value = label
        ws.cell(row=header_row, column=2).value = metric_label
        for c in (1, 2):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(bold=True, color=header_font_color)
            cell.fill = PatternFill(fill_type="solid", fgColor=header_fill_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        start_data_row = header_row + 1
        for i, (k, v) in enumerate(totals.items(), start=start_data_row):
            ws.cell(row=i, column=1).value = k
            ws.cell(row=i, column=2).value = float(_agg_result(v, agg))
            if stripe_fill is not None and ((i - start_data_row) % 2 == 1):
                ws.cell(row=i, column=1).fill = stripe_fill
                ws.cell(row=i, column=2).fill = stripe_fill
        ws.freeze_panes = ws.cell(row=start_data_row, column=1)
        if totals:
            ws.auto_filter.ref = f"A{header_row}:B{start_data_row + len(totals) - 1}"

    max_cols = max(2, int(getattr(ws, "max_column", 2) or 2))
    for i in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14 if i > 1 else 22


def _add_pivot_dynamic_sheet(
    wb: Workbook,
    columns: list[str],
    rows: list[list[Any]],
    pivot_config: dict[str, Any] | None = None,
) -> None:
    if not columns or not rows:
        return

    cfg = _normalize_pivot_config(pivot_config)
    numeric_idxs = [idx for idx in range(len(columns)) if _is_numeric_column(rows, idx)]
    text_idxs = [idx for idx in range(len(columns)) if idx not in numeric_idxs]

    default_row = (cfg.get("rows") or [columns[text_idxs[0] if text_idxs else 0]])[0]
    if str(default_row or "").strip() not in columns:
        default_row = columns[text_idxs[0] if text_idxs else 0]

    default_col = (cfg.get("columns") or [columns[text_idxs[1]] if len(text_idxs) > 1 else (columns[text_idxs[0]] if text_idxs else columns[0])])[0]
    if str(default_col or "").strip() not in columns:
        default_col = columns[text_idxs[1]] if len(text_idxs) > 1 else (columns[text_idxs[0]] if text_idxs else columns[0])

    default_val = str(cfg.get("value") or "").strip()
    if default_val not in columns:
        if numeric_idxs:
            default_val = columns[numeric_idxs[0]]
        else:
            default_val = columns[0]

    default_agg = str(cfg.get("agg") or "sum").strip().upper()
    if default_agg not in {"SUM", "COUNT", "AVERAGE", "MIN", "MAX"}:
        default_agg = "SUM"

    ws = wb.create_sheet(title="PivotDynamic")
    ws["A1"].value = "Dynamic pivot (Excel 365)"
    ws["A1"].font = Font(size=13, bold=True)
    ws["A2"].value = "Select fields below. The matrix updates automatically."

    ws["A4"].value = "Row field"
    ws["C4"].value = "Column field"
    ws["E4"].value = "Value field"
    ws["G4"].value = "Aggregation"
    for cell in ("A4", "C4", "E4", "G4"):
        ws[cell].font = Font(bold=True)

    ws["A5"].value = str(default_row)
    ws["C5"].value = str(default_col)
    ws["E5"].value = str(default_val)
    ws["G5"].value = str(default_agg)

    # Header list for dropdowns.
    ws["J3"].value = "Fields"
    ws["J3"].font = Font(bold=True)
    for idx, col in enumerate(columns, start=4):
        ws.cell(row=idx, column=10).value = str(col)

    data_first_row = 4
    data_last_row = 3 + len(rows)
    data_last_col = get_column_letter(len(columns))
    data_range = f"Data!$A${data_first_row}:${data_last_col}${data_last_row}"
    header_range = f"Data!$A$3:${data_last_col}$3"
    fields_ref = f"$J$4:$J${3 + len(columns)}"

    dv_fields = DataValidation(type="list", formula1=fields_ref, allow_blank=False)
    ws.add_data_validation(dv_fields)
    dv_fields.add(ws["A5"])
    dv_fields.add(ws["C5"])
    dv_fields.add(ws["E5"])

    dv_agg = DataValidation(type="list", formula1='"SUM,COUNT,AVERAGE,MIN,MAX"', allow_blank=False)
    ws.add_data_validation(dv_agg)
    dv_agg.add(ws["G5"])

    ws["A7"].value = "Rows"
    ws["B7"].value = "Columns"
    ws["A7"].font = Font(bold=True)
    ws["B7"].font = Font(bold=True)

    # Dynamic arrays for row and column members.
    ws["A8"].value = (
        f"=SORT(UNIQUE(FILTER(INDEX({data_range},0,MATCH($A$5,{header_range},0)),"
        f"INDEX({data_range},0,MATCH($A$5,{header_range},0))<>\"\")))"
    )
    ws["B7"].value = (
        f"=TRANSPOSE(SORT(UNIQUE(FILTER(INDEX({data_range},0,MATCH($C$5,{header_range},0)),"
        f"INDEX({data_range},0,MATCH($C$5,{header_range},0))<>\"\"))))"
    )

    # Build matrix with MAKEARRAY/LAMBDA (Excel 365).
    ws["B8"].value = (
        f"=MAKEARRAY(ROWS($A$8#),COLUMNS($B$7#),"
        f"LAMBDA(r,c,LET(rv,INDEX($A$8#,r),cv,INDEX($B$7#,c),"
        f"vals,INDEX({data_range},0,MATCH($E$5,{header_range},0)),"
        f"rvals,INDEX({data_range},0,MATCH($A$5,{header_range},0)),"
        f"cvals,INDEX({data_range},0,MATCH($C$5,{header_range},0)),"
        f"IF($G$5=\"COUNT\",COUNTIFS(rvals,rv,cvals,cv),"
        f"IF($G$5=\"AVERAGE\",AVERAGEIFS(vals,rvals,rv,cvals,cv),"
        f"IF($G$5=\"MIN\",MINIFS(vals,rvals,rv,cvals,cv),"
        f"IF($G$5=\"MAX\",MAXIFS(vals,rvals,rv,cvals,cv),"
        f"SUMIFS(vals,rvals,rv,cvals,cv))))))))"
    )

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["J"].width = 24
    ws.freeze_panes = "A8"


def table_to_xlsx_bytes(
    title: str,
    columns: list[str],
    rows: list[list[Any]],
    *,
    add_chart: bool = True,
    chart_title: str | None = None,
    top_n: int = 15,
    template: str = "clean",
    color_theme: str | None = None,
    add_pivot: bool = False,
    pivot_config: dict[str, Any] | None = None,
) -> bytes:
    """Create an .xlsx file from a tabular dataset.

    Chart behavior (best effort):
    - If there are at least 2 columns, and the 2nd column is mostly numeric,
      create a bar chart using col1 as category and col2 as value.
    """
    template_key = (template or "clean").strip().lower()
    style = _TEMPLATE_STYLES.get(template_key, _TEMPLATE_STYLES["clean"])

    header_fill_color = _resolve_theme_color(color_theme) or style["header_fill"]
    header_font_color = style["header_font"]
    stripe_fill_color = style.get("stripe_fill")
    title_fill_color = style.get("title_fill")
    title_font_color = style.get("title_font")
    chart_style = int(style.get("chart_style") or 10)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Title row (optional, but nice)
    title = (title or "Export").strip() or "Export"
    ws["A1"].value = title
    ws["A1"].font = Font(size=14, bold=True, color=title_font_color)
    if title_fill_color:
        ws["A1"].fill = PatternFill(fill_type="solid", fgColor=title_fill_color)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Header
    header_row = 3
    for c_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx)
        cell.value = str(col)
        cell.font = Font(bold=True, color=header_font_color)
        cell.fill = PatternFill(fill_type="solid", fgColor=header_fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    max_rows = 200000  # practical safety
    rows = rows[:max_rows]
    apply_stripe = bool(stripe_fill_color and len(rows) <= 15000 and len(columns) <= 50)
    stripe_fill = PatternFill(fill_type="solid", fgColor=stripe_fill_color) if apply_stripe else None
    for r_idx, r in enumerate(rows, start=header_row + 1):
        is_striped = apply_stripe and ((r_idx - (header_row + 1)) % 2 == 1)
        for c_idx, v in enumerate(r[: len(columns)], start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = _safe_cell(v)
            if is_striped and stripe_fill is not None:
                cell.fill = stripe_fill

    # Freeze + filter
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if columns:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(rows)}"

    # Column widths (best effort)
    for i, col in enumerate(columns, start=1):
        width = min(45, max(10, len(str(col)) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width

    # Optional chart
    if add_chart and len(columns) >= 2 and rows:
        # Determine if 2nd column is numeric enough
        sample = rows[: min(50, len(rows))]
        numeric_hits = sum(1 for r in sample if len(r) >= 2 and _is_number(r[1]))
        if numeric_hits >= max(5, int(0.6 * len(sample))):
            ws_chart = wb.create_sheet(title="Chart")
            ws_chart["A1"].value = chart_title or "Chart"
            ws_chart["A1"].font = Font(size=14, bold=True)

            # Copy top N (sorted by value desc) to chart sheet for stable references
            data_pairs: list[tuple[Any, float]] = []
            for r in rows:
                if len(r) < 2:
                    continue
                if _is_number(r[1]):
                    try:
                        data_pairs.append((r[0], float(r[1])))
                    except Exception:
                        pass
            data_pairs.sort(key=lambda x: x[1], reverse=True)
            data_pairs = data_pairs[: max(1, int(top_n))]

            ws_chart["A3"].value = str(columns[0])
            ws_chart["B3"].value = str(columns[1])
            ws_chart["A3"].font = Font(bold=True)
            ws_chart["B3"].font = Font(bold=True)

            for i, (cat, val) in enumerate(data_pairs, start=4):
                ws_chart[f"A{i}"].value = _safe_cell(cat)
                ws_chart[f"B{i}"].value = val

            chart = BarChart()
            chart.type = "col"
            chart.style = chart_style
            chart.title = chart_title or f"{columns[1]} by {columns[0]}"
            chart.y_axis.title = str(columns[1])
            chart.x_axis.title = str(columns[0])

            data_ref = Reference(ws_chart, min_col=2, min_row=3, max_row=3 + len(data_pairs))
            cats_ref = Reference(ws_chart, min_col=1, min_row=4, max_row=3 + len(data_pairs))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 13
            chart.width = 26
            ws_chart.add_chart(chart, "D3")

    if add_pivot and len(columns) >= 2 and rows:
        _add_pivot_sheet(
            wb,
            title,
            columns,
            rows,
            header_fill_color=header_fill_color,
            header_font_color=header_font_color,
            stripe_fill_color=stripe_fill_color,
            pivot_config=pivot_config,
        )
        _add_pivot_dynamic_sheet(
            wb,
            columns,
            rows,
            pivot_config=pivot_config,
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
